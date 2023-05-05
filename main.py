"""
What I want to do:
- create a class with attributes for each herb, recording: herb name, shortened herb name, death rate, number of
harvests, deaths, average yield, median yield and list of harvest yields
    - not recording # of farm runs bc a run can have multiple types of herbs (this make so sense if all herbs have same
    death rate and yield rate
- class has a function that calculates the class attributes after info has been recorded
- harvest data recorded into a spreadsheet for ease of use then imported to each object through a class method

this is inefficient at a large scale but it's fine for this small project
 """

import global_data as gd
from generate_default_spreadsheet import create_default_spreadsheet
from statistics import median
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class HerbTable:

    def __init__(self, name: str):
        self.name = name
        self.harvest_yields = []

        self.deaths = 0
        self.nr_harvests = 0
        self.lowest_yield = None
        self.highest_yield = 0

        self.death_rate = None
        self.total_harvested = None
        self.average_yield = None
        self.median_yield = None
        self.yield_per_seed = None

    def calc_class_attrs(self):
        self.death_rate = self.deaths / self.nr_harvests
        self.total_harvested = sum(self.harvest_yields)
        self.average_yield = self.total_harvested / len(self.harvest_yields)
        self.median_yield = median(self.harvest_yields)
        self.yield_per_seed = self.average_yield * 1.1

    def add_yields(self, yields: list):
        for yy in yields:
            self.nr_harvests += 1
            if yy == 0:
                self.deaths += 1
            elif self.lowest_yield is None or yy < self.lowest_yield:
                self.lowest_yield = yy
            elif yy > self.highest_yield:
                self.highest_yield = yy
            self.harvest_yields.append(yy)

    def print_stats(self):
        if self.lowest_yield is None:
            return False
        print(f"""
        Stats for {self.name} harvests:
        
        # of Harvests:   {self.nr_harvests}
        Total Harvested: {self.total_harvested}
        Lowest Yield:    {self.lowest_yield}
        Highest Yield:   {self.highest_yield}
        
        Death Chance:    {self.death_rate * 100}%
        Median Yield:    {self.median_yield}
        Average Yield:   {self.average_yield}
        Yield per Seed:  {self.yield_per_seed}
""")
        return True

    def validate_data(self):
        # bug testing function
        if len(self.lowest_yield) is None:
            return False
        print(f"\nValidating data for {self.name}...")
        if self.harvest_yields.count(0) != self.deaths:
            print("NOT OK: Deaths do not match.")
        else:
            print("OK: Deaths match.")
        if self.nr_harvests != len(self.harvest_yields):
            print("NOT OK: Number of harvests do NOT match.")
        else:
            print("OK: Number of harvests match.")
        print(
            f"Check: average_yield = {self.average_yield} | average = {sum(self.harvest_yields) / len(self.harvest_yields)}")
        print('Done validating.')
        return True


def calc_cumulative_data_to_yields(yield_list: list):
    """transforms cumulative data into yield per allotment harvest"""
    first_yield = yield_list[0]
    real_yields = [first_yield]
    total_yield = first_yield
    for yy in yield_list[1:]:
        real_yields.append(yy - total_yield)
        total_yield = yy
    return real_yields


def update_herb_objects_through_harvest_sheet(inp_sheet, herb_objs_dict: dict, herb_column_idxs: list):
    green_cell_ranges = []
    # herb columns loop
    for hcidx in herb_column_idxs:
        current_herb_name = inp_sheet.cell(row=1, column=hcidx).value
        cumulative_yields_list = []
        yields_list = []

        current_row = 2
        current_column = hcidx
        found_first_green_cell = False
        green_cell_range_begin = ""
        green_cell_range_end = ""

        # rows loop
        while True:
            current_row_cell = inp_sheet.cell(row=current_row, column=hcidx)
            # if the current row is empty, break the infinite rows loop, going to the next herb - and update last green
            # cell
            if current_row_cell.value is None:
                green_cell_range_end = inp_sheet.cell(row=current_row - 1, column=current_column + 4).coordinate
                break
            # if green cell, skip
            elif current_row_cell.fill.start_color.index == gd.finished_cell_green:
                current_row += 1
                continue
            else:
                if not found_first_green_cell:
                    green_cell_range_begin = current_row_cell.coordinate
                    found_first_green_cell = True
                # loop through the row
                for cc in range(hcidx, hcidx + 6):
                    cell_value = inp_sheet.cell(row=current_row, column=current_column).value
                    if type(cell_value) is int and cell_value >= 0:
                        cumulative_yields_list.append(cell_value)
                        current_column += 1
                    elif cell_value is None or cc == hcidx + 5:
                        yields_list += calc_cumulative_data_to_yields(cumulative_yields_list)
                        cumulative_yields_list = []
                        # go to next row and reset column index
                        current_row += 1
                        current_column = hcidx
                        break
        if found_first_green_cell:
            green_cell_ranges.append([green_cell_range_begin, green_cell_range_end])
        if yields_list:
            herb_objs_dict[current_herb_name].add_yields(yields_list)
            herb_objs_dict[current_herb_name].calc_class_attrs()
    return herb_objs_dict, green_cell_ranges


def convert_HerbTables_to_stats_sheet(herb_objs_dict: dict, sts_sheet):
    for rr in range(2, len(gd.herb_names) + 2):
        herb_name = sts_sheet.cell(row=rr, column=1).value
        sts_sheet.cell(row=rr, column=2).value = herb_objs_dict[herb_name].nr_harvests
        sts_sheet.cell(row=rr, column=3).value = herb_objs_dict[herb_name].total_harvested
        sts_sheet.cell(row=rr, column=4).value = herb_objs_dict[herb_name].lowest_yield
        sts_sheet.cell(row=rr, column=5).value = herb_objs_dict[herb_name].highest_yield
        sts_sheet.cell(row=rr, column=6).value = herb_objs_dict[herb_name].death_rate
        sts_sheet.cell(row=rr, column=7).value = herb_objs_dict[herb_name].median_yield
        sts_sheet.cell(row=rr, column=8).value = herb_objs_dict[herb_name].average_yield
        sts_sheet.cell(row=rr, column=9).value = herb_objs_dict[herb_name].yield_per_seed
        harvest_yields_string = ", ".join([str(hh) for hh in herb_objs_dict[herb_name].harvest_yields])
        sts_sheet.cell(row=rr, column=10).value = harvest_yields_string


def convert_stats_sheet_to_HerbTables(sts_sheet, herb_objs_dict: dict):
    for rr in range(2, len(gd.herb_names) + 2):
        herb_name = sts_sheet.cell(row=rr, column=1).value
        if not sts_sheet.cell(row=rr, column=2).value:
            continue
        herb_objs_dict[herb_name].nr_harvests = int(sts_sheet.cell(row=rr, column=2).value)
        herb_objs_dict[herb_name].total_harvested = int(sts_sheet.cell(row=rr, column=3).value)
        herb_objs_dict[herb_name].lowest_yield = int(sts_sheet.cell(row=rr, column=4).value)
        herb_objs_dict[herb_name].highest_yield = int(sts_sheet.cell(row=rr, column=5).value)
        herb_objs_dict[herb_name].death_rate = float(sts_sheet.cell(row=rr, column=6).value)
        herb_objs_dict[herb_name].median_yield = int(sts_sheet.cell(row=rr, column=7).value)
        herb_objs_dict[herb_name].average_yield = float(sts_sheet.cell(row=rr, column=8).value)
        herb_objs_dict[herb_name].yield_per_seed = float(sts_sheet.cell(row=rr, column=9).value)
        harvest_yields_list = [int(ll) for ll in sts_sheet.cell(row=rr, column=10).value.split(", ")]
        herb_objs_dict[herb_name].harvest_yields = harvest_yields_list
    return herb_objs_dict


def print_all_stats_to_console(herb_objs_dict: dict):
    empty_herbs = []
    for herb_name, herb_objct in herb_objs_dict.items():
        if not herb_objct.nr_harvests:
            empty_herbs.append(herb_name)
        else:
            herb_objct.print_stats()
    print(f"Herbs with empty data: {empty_herbs}")


if __name__ == '__main__':
    herb_book = None
    try:
        herb_book = load_workbook(gd.spreadsheet_path)
    except FileNotFoundError:
        print(f"{gd.spreadsheet_path} does not exist.")
        print(f"Create spreadsheet? [y/n]")
        if input().lower() == 'y':
            create_default_spreadsheet()
        exit()

    input_sheet = herb_book[gd.herbsheet_names[0]]
    stats_sheet = herb_book[gd.herbsheet_names[1]]

    # todo: working now - how to update objects if the class is updated?
    # create herb objects into a dictionary and import data into them through the spreadsheet
    herb_idx_locs_in_first_sheet = list(range(2, 3 + ((len(gd.herb_names) - 1) * 6), 6))
    herb_objects = {herb_name: HerbTable(herb_name) for herb_name in gd.herb_names}
    herb_objects = convert_stats_sheet_to_HerbTables(stats_sheet, herb_objects)
    while True:
        print("[add] new data from spreadsheet, show [stats], [exit]")
        match input().lower():

            case "add":
                # import input sheet data into objects
                herb_objects, g_c_ranges = update_herb_objects_through_harvest_sheet(input_sheet, herb_objects,
                                                                                     herb_idx_locs_in_first_sheet)
                # color finished cells green
                for cell_ranges in g_c_ranges:
                    cell_range = input_sheet[cell_ranges[0]:cell_ranges[1]]
                    for cell_tuple in cell_range:
                        for cell in cell_tuple:
                            cell.fill = PatternFill("solid", fgColor=gd.finished_cell_green)

                # update stats sheet
                convert_HerbTables_to_stats_sheet(herb_objects, stats_sheet)
                herb_book.save(gd.spreadsheet_path)
                print("Data from spreadsheet parsed.")
                print_all_stats_to_console(herb_objects)
                exit()

            case "stats":
                print_all_stats_to_console(herb_objects)
                exit()

            case "exit":
                print("Goodbye!")
                exit()

            case _:
                "Invalid input."
    pass
