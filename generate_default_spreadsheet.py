from openpyxl import Workbook
from openpyxl.styles import PatternFill
from global_data import spreadsheet_name, spreadsheet_path, herb_names, herbsheet_names, stats_name_list


def create_default_spreadsheet():
    herb_book = Workbook()
    harvest_input_sheet = herb_book.active
    harvest_input_sheet.title = herbsheet_names[0]
    stats_sheet = herb_book.create_sheet(herbsheet_names[1])

    # input data sheet
    harvest_input_sheet.cell(row=1, column=1).value = "Herbs:"
    harvest_input_sheet.cell(row=2, column=1).value = "Harvests:"
    col_range = list(range(2, 3 + ((len(herb_names) - 1) * 6), 6))
    for yy, herb_name in zip(col_range, herb_names):
        this_cell = harvest_input_sheet.cell(row=1, column=yy)
        this_cell.value = herb_name
        this_cell.fill = PatternFill("solid", fgColor="00008000")
        pass

    # stats sheet
    for yy, stat_name in enumerate(stats_name_list, 2):
        stats_sheet.cell(row=1, column=yy).value = stat_name
    # brown the yields list column?
    for xx, herb_name in enumerate(herb_names, 2):
        stats_sheet.cell(row=xx, column=1).value = herb_name

    print(f"{spreadsheet_name} created.")
    herb_book.save(spreadsheet_path)
